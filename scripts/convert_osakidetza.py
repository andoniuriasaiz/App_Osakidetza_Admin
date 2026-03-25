#!/usr/bin/env python3
import openpyxl
import json
import os
import re
from pathlib import Path

SOURCE_DIR = "/Users/andoniuria/Library/CloudStorage/GoogleDrive-uriasaiz@gmail.com/Mi unidad/40-49_Trabajo/44_Oposiciones/04_OSAKIDETZA/20_BATERIA_PREGUNTAS/30_VAIA_IMPORT/REVISADAS PARA IMPORTAR"
TARGET_DIR = "/Users/andoniuria/Library/CloudStorage/GoogleDrive-uriasaiz@gmail.com/Mi unidad/40-49_Trabajo/44_Oposiciones/04_OSAKIDETZA/60_HERRAMIENTAS/App_Osakidetza_Admin/public/data"

def get_module_id(filename):
    # COM_C2_T01_Profesiones_Sanitarias.xlsx -> comun-t01
    # AUX_E01_EAPV_Estatuto_Autonomia.xlsx -> aux-e01
    # ADM_EBEP_Empleado_Publico.xlsx -> admin-ebep
    
    stem = Path(filename).stem
    if stem.startswith("COM_"):
        match = re.search(r"T(\d+)", stem)
        if match:
            return f"comun-t{match.group(1)}"
        return stem.lower().replace("_", "-")
    elif stem.startswith("AUX_"):
        match = re.search(r"E(\d+)", stem)
        if match:
            return f"aux-e{match.group(1)}"
        return stem.lower().replace("_", "-")
    elif stem.startswith("ADM_"):
        return stem.lower().replace("_", "-")
    return stem.lower().replace("_", "-")

def process_file(filepath, module_id):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    questions = []
    headers = [cell.value for cell in ws[1]]
    
    # Headers expected: 0: Question, 1: AnsA, 2: CorrectA, 3: AnsB, 4: CorrectB... 13: Tags, 14: Hints, 15: Explanation
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        if not row[0]: continue
        
        question_text = row[0]
        options = []
        correct_answer_nums = []
        correct_answers = []
        
        # Options A to F are in columns 1, 3, 5, 7, 9, 11
        # Correctness is in 2, 4, 6, 8, 10, 12
        for i in range(6):
            ans_idx = 1 + (i * 2)
            cor_idx = 2 + (i * 2)
            if ans_idx < len(row) and row[ans_idx] is not None:
                val = i + 1
                opt_text = str(row[ans_idx]).strip()
                options.append({"value": val, "text": opt_text})
                
                # Check if correct (TRUE/True/1 or if it's explicitly 'TRUE')
                is_correct = row[cor_idx]
                if is_correct is True or str(is_correct).upper() == "TRUE":
                    correct_answer_nums.append(val)
                    correct_answers.append(opt_text)
        
        explanation = row[15] if len(row) > 15 else None
        tags = row[13] if len(row) > 13 else None
        
        questions.append({
            "id": f"{module_id}_{row_idx}",
            "questionNum": row_idx,
            "question": str(question_text).strip(),
            "type": "C", # Multiple choice
            "options": options,
            "correctAnswerNums": correct_answer_nums,
            "correctAnswers": correct_answers,
            "multipleCorrect": len(correct_answer_nums) > 1,
            "hasImage": False,
            "image": None,
            "imageUrl": None,
            "module": module_id,
            "explanation": explanation
        })
    
    return questions

def main():
    os.makedirs(TARGET_DIR, exist_ok=True)
    all_files = [f for f in os.listdir(SOURCE_DIR) if f.endswith(".xlsx")]
    
    for filename in all_files:
        module_id = get_module_id(filename)
        filepath = os.path.join(SOURCE_DIR, filename)
        
        print(f"Processing {filename} -> {module_id}")
        try:
            questions = process_file(filepath, module_id)
            target_file = os.path.join(TARGET_DIR, f"{module_id}.json")
            with open(target_file, "w", encoding="utf-8") as f:
                json.dump(questions, f, ensure_ascii=False, indent=2)
            print(f"  Done: {len(questions)} questions")
        except Exception as e:
            print(f"  Error processing {filename}: {e}")

if __name__ == "__main__":
    main()
